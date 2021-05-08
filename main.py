from constants import *
from driver_paths import *
from selenium import webdriver
import openpyxl
import xlsxwriter
import os
import datetime

# Open link
driver = webdriver.Chrome(executable_path=path_chromedriver)
driver.get('https://ogero.gov.lb/myogero/consumption.php')

# Login to Ogero
username = driver.find_element_by_xpath(username_path)
password = driver.find_element_by_xpath(password_path)
submit_button = driver.find_element_by_xpath(submit_path)
username.send_keys(OgeroUsername)
password.send_keys(OgeroPassword)
submit_button.click()

# Extract relevant information
upload = driver.find_element_by_xpath(upload_path).text
download = driver.find_element_by_xpath(download_path).text
total = driver.find_element_by_xpath(total_path).text
extra_consumption = driver.find_element_by_xpath(extra_consumption_path).text
consumption_until = driver.find_element_by_xpath(consumption_until_path).text

excel_path = "My-Internet-Consumption.xlsx"
# If you're running the script for the first time, create the Excel file with appropriate formatting
if not os.path.exists(excel_path):
    workbook = xlsxwriter.Workbook(excel_path)
    worksheet = workbook.add_worksheet()
    bold = workbook.add_format({'bold': 1})
    worksheet.write('A1', "Date", bold)
    worksheet.write('B1', "Upload", bold)
    worksheet.write('C1', "Download", bold)
    worksheet.write('D1', "Total", bold)
    worksheet.write('E1', "Extra Consumption", bold)
    worksheet.write('F1', "Date Updated", bold)
    workbook.close()

# Add the info to existing Excel file
wb_obj = openpyxl.load_workbook(excel_path)
sheet_obj = wb_obj.active
row_update = sheet_obj.max_row + 1

# Minor check here: make sure this is not a duplicate entry
for row in range(1, row_update):
    if sheet_obj.cell(row=row, column=1).value == consumption_until and \
            sheet_obj.cell(row=row, column=2).value == upload and \
            sheet_obj.cell(row=row, column=3).value == download and \
            sheet_obj.cell(row=row, column=4).value == total and \
            sheet_obj.cell(row=row, column=5).value == extra_consumption:
        break
    else:
        sheet_obj.cell(row=row_update, column=1, value=consumption_until)
        sheet_obj.cell(row=row_update, column=2, value=upload)
        sheet_obj.cell(row=row_update, column=3, value=download)
        sheet_obj.cell(row=row_update, column=4, value=total)
        sheet_obj.cell(row=row_update, column=5, value=float(extra_consumption))
        sheet_obj.cell(row=row_update, column=6, value=datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        wb_obj.save(excel_path)
